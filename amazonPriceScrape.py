import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

import tkinter
from tkinter import messagebox
import win32com.client
from win32com.client.gencache import EnsureDispatch
from threading import Thread
from multiprocessing import Process
import os
import sys
from selenium.common.exceptions import NoSuchElementException
sys.path.append("G:\\Ranjan\\2020\\29-01-2020\\Price_Scrape_Automation")
sys.path.append("G:\\Ranjan\\2020\\29-01-2020\\Price_Scrape_Automation\\engine")
from engine.app.setup import config





def showInfo(infoMessage,duration,browser):	
	try:

		jscode = '''alert('Collecting the prices and other details from the amazon website for the Search Item')'''
		browser.execute_script(jscode)
		alert = browser.switch_to.alert
		time.sleep(6)
		#print("Open alert in item search functin")
		alert.accept()
	except Exception as e:
		print(str(e)+"alert exception")
		pass




def highlight(element):
    """Highlights a Selenium webdriver element"""
    driver = element._parent
    def apply_style(s):
        driver.execute_script("arguments[0].setAttribute('style', arguments[1])", element, s)
    orignal_style = element.get_attribute('style')
    apply_style("background-color:red")
    if (element.get_attribute("style")!=None):
        time.sleep(1)
    
    apply_style(orignal_style)


def classSyntax(className):
	className=className.replace(" ",".")
	return className



def itemSearch(browser,itemName,configSetup):
	
	qryData1=configSetup["searchQry"]["class_value"][0]
	qryData2=configSetup["resElement"]["class_value"][0]

	showInfo("Collecting the prices and other details from the amazon website for the %s"%itemName,5000,browser)

	browser.find_element_by_id(qryData1).send_keys(itemName)
	browser.find_element_by_id(qryData1).send_keys(u'\ue007')


	search_1=browser.find_elements_by_class_name(classSyntax(qryData2))
	for ls in search_1:
		try:
			lst=getAllchildern=ls.find_elements_by_xpath(".//*")
			if len(lst)!=0:
				search_1=ls
		except Exception as e:
			print(e,"if conditions for elements")
			pass


	#search_1=search_1[1]
	getAllchildern=search_1.find_elements_by_xpath(".//*")
	searchClass=getAllchildern[0].get_attribute('class')
	searchClass=classSyntax(searchClass)
	retSearch=search_1.find_elements_by_class_name(classSyntax(searchClass))
	

	return retSearch



def nameFun(search_1,browser,configSetup):	
	nameVal=''

	tagQuery1=configSetup["nameQry"]["tag_value"][0]
	try:
		elements=search_1.find_elements_by_css_selector(tagQuery1)
		#print(elements[0].get_attribute("innerHTML"))		
		if not elements:
		    print("No element found")
		    nameVal="not found"  
		else:
			name1=search_1.find_element_by_css_selector(tagQuery1)
			highlight(name1)
			nameVal=name1.get_attribute('innerHTML')  
	except NoSuchElementException as e:
		print("NoSuchElementException#######################-NAME",str(e))
	return nameVal




def ratingFun(search_1,browser,configSetup):		
	ratingVal=''
	classQry1=classSyntax(configSetup["ratingQry"]["class_value"][0])
	tagQuery1=configSetup["ratingQry"]["tag_value"][0]
	tagQueryHighlight=tagQuery1.split()
	tagQueryHighlight.pop(-1)
	tagQueryHighlight=' '.join(tagQueryHighlight)
	#print(tagQueryHighlight,"tagQueryHighlight")
	

	try:
		elements=search_1.find_elements_by_css_selector(tagQuery1)

		if not elements:
		    print("No element found rating")  
		    ratingVal="not found"
		else:
			ratings1=search_1.find_element_by_css_selector(tagQuery1)
			ratings1HIGH=search_1.find_element_by_css_selector(tagQueryHighlight)
			highlight(ratings1HIGH)
			ratingVal=ratings1.get_attribute('innerHTML')     
	except NoSuchElementException as e:
		print(str(e),"rating val function exception")
	return ratingVal






def oriPriceFun(search_1,browser,configSetup):
	oriPriceVal=''
	classQry1=classSyntax(configSetup["oriPriceQry"]["class_value"][0])
	classQry2=classSyntax(configSetup["oriPriceQry"]["class_value"][1])
	try:		
		elements=search_1.find_elements_by_class_name(classQry1)

		if not elements:
		    print("No element found oriPriceVal")  
		    oriPriceVal="Not Found"
		else:
			oriPrice1=search_1.find_element_by_class_name(classQry1)
			oriPrice2=oriPrice1.find_element_by_class_name(classQry2)
			oriPriceVal=oriPrice2.get_attribute('innerHTML')  
			highlight(oriPrice1)
	except NoSuchElementException as e:
		print(str(e)," oriPriceVal function ")
	return oriPriceVal
	



def curPriceFun(search_1,browser,configSetup):
	curPriceVal=''
	classQry1=classSyntax(configSetup["curPriceQry"]["class_value"][0])
	try:		
		elements=search_1.find_elements_by_class_name(classQry1)
		if not elements:
		    print("No element found curPriceVal")  
		    curPriceVal="not found"
		else:
			curPrice1=search_1.find_element_by_class_name(classQry1)
			highlight(curPrice1)
			curPriceVal=curPrice1.get_attribute('innerHTML')
	except NoSuchElementException as e:
		print(str(e),"curPriceVal function exception")
	return curPriceVal
		
	



def shippDateFun(search_1,browser,configSetup):
	shippDateVal=''
	tagQuery1=str(configSetup["shipDateQry"]["tag_value"][0])
	try:
		elements=search_1.find_elements_by_css_selector(tagQuery1)			
		if not elements:
		    print("No element found shippDateVal")  
		    shippDateVal="no element"
		else:
			shippDate1=search_1.find_element_by_css_selector(tagQuery1)
			highlight(shippDate1)
			shippDateVal=shippDate1.get_attribute('innerHTML')
	except NoSuchElementException as e:
		print(str(e),"shippDateVal function exception")
	return shippDateVal



def writeheaders(opFile):
	opWorkbook=openpyxl.Workbook()
	opWorksheet=opWorkbook.active
	opWorksheet.title = "itemDetails"
	opWorksheet.column_dimensions['A'].width = 30
	opWorksheet.column_dimensions['B'].width = 25
	opWorksheet.column_dimensions['C'].width = 20
	opWorksheet.column_dimensions['D'].width = 20
	opWorksheet.column_dimensions['E'].width = 20
	opWorksheet.column_dimensions['F'].width = 20
	
	opWorksheet['A1']= "Names"
	opWorksheet['B1']= "Ratings"				
	opWorksheet['C1']= "Original Price"
	opWorksheet['D1']= "Current Price"
	opWorksheet['E1']= "Shipping Date"
	opWorksheet['F1']= "Source Website"

	opWorkbook.save(filename=opFile)




def writeExcelFun(nameVal,ratingVal,oriPriceVal,curPriceVal,shippDateVal,url,opFile,numItr,cellNum,browser):

	try:

		print("actual writing",nameVal,cellNum,type(cellNum))
		opWorkbook=openpyxl.load_workbook(opFile)
		opWorksheet=opWorkbook.worksheets[0]
		num=cellNum+2
		opWorksheet['A%s'%num]=nameVal
		opWorksheet['B%s'%num]=ratingVal
		opWorksheet['C%s'%num]=oriPriceVal
		opWorksheet['D%s'%num]=curPriceVal
		opWorksheet['E%s'%num]=shippDateVal
		opWorksheet['F%s'%num]=url
		opWorkbook.save(filename=opFile)
	except Exception as e:
		print(e,"alert exceptions")
	

def openExcel(opFile):
	xl = win32com.client.Dispatch("Excel.Application")
	xl.Visible = True # otherwise excel is hidden
	wb = xl.Workbooks.Open(opFile)
	#xl.WindowState = win32.constants.xlMaximized
	# time.sleep(10)
	# wb.Close()
	# xl.Quit()



def startFunc(ipFile,opFile,chromedriver_path,selected_website):

	start_time = time.time()

	# remove op file if already exists
	if os.path.isfile(opFile):
		os.remove(opFile)	

	if selected_website=="Amazon":
		url="https://www.amazon.in/"


	# read the config file	
	confKeys=list(config.keys())
	configSetup=config
	

	
	# read the input from the input excel file
	wb_ip = openpyxl.load_workbook(ipFile)
	ip_sheet = wb_ip.worksheets[0]
	itemName=str(ip_sheet[2][1].value)
	numItr=int(ip_sheet[2][2].value)


	# check if file exists and if or if not write headers
	writeheaders(opFile)



	
	# open the browser , maximize, and search
	chrome_driver_path = chromedriver_path
	opts = Options()	
	#opts.add_argument("--headless")
	browser = webdriver.Chrome(executable_path=chrome_driver_path, options=opts)
	browser.maximize_window()
	url=str(url)	
	browser.get(url)

	try:
		jscode = '''   alert("Please wait while the Bot is searching for the required Details of the Item")   '''
		browser.execute_script(jscode)
		alert = browser.switch_to.alert			
		alert.accept()	
	except Exception as e:
		print(str(e)+"alert exception at the start program")
		pass


	# search the item and return the element
	search_1=itemSearch(browser,itemName,configSetup)






	for i in range(numItr):
		try:
			browser.execute_script("return arguments[0].scrollIntoView();", search_1[i])
			browser.execute_script("window.scrollBy(0, -150);")
		except Exception as e:
			print(str(e),"321 for focus")

		nameVal=nameFun(search_1[i],browser,configSetup)	
		ratingVal=ratingFun(search_1[i],browser,configSetup)
		curPriceVal=curPriceFun(search_1[i],browser,configSetup)
		oriPriceVal=oriPriceFun(search_1[i],browser,configSetup)
		shippDateVal=shippDateFun(search_1[i],browser,configSetup)
		
		try:	
			jscode = '''   alert("Please wait the excel is been updating")   '''
			browser.execute_script(jscode)
			alert = browser.switch_to.alert
			cellNum=i
			writeExcelFun(nameVal,ratingVal,oriPriceVal,curPriceVal,shippDateVal,url,opFile,numItr,cellNum,browser)
			alert.accept()
		except Exception as e:
			print(e,"exception handling at writeExcelFun")

	try:
		jscode = '''   alert("Finished the process of getting the product info now the crome browser will be closed and the Output Excel generated file will open")   '''
		browser.execute_script(jscode)
		alert = browser.switch_to.alert
		openExcel(opFile)	
		alert.accept()	
	except Exception as e:
		print(str(e)+"alert exception at the start program")		
	
	browser.close()
	browser.quit()

		



if __name__ == '__main__':
	startFunc("G:\\Ranjan\\2020\\jan_2020\\input_data.xlsx","G:\\Ranjan\\2020\\jan_2020\\Product-info-File.xlsx",'G:\\Ranjan\\2020\\29-01-2020\\Price_Scrape_Automation\\bin\\selenium-drivers\\Chrome\\75\\chromedriver.exe',"Amazon")





